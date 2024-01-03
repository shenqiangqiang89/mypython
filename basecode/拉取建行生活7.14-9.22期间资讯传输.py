# -*- coding: utf-8 -*-

from datetime import datetime, timedelta
import requests
from openpyxl import Workbook

# 创建 Workbook （工作簿）和 Worksheet （工作表）
wb = Workbook()
ws = wb.active  # 获取当前工作表
# 写入数据
ws['A1'] = '更新时间'
ws['B1'] = '内容类型'
ws['C1'] = '内容分类'
ws['D1'] = '标题'

url = 'https://w-1.betawm.com/betanewsextapi/Content/GetList?AliAppId=110986569'
headers = {
    "Content-Type": "application/json; charset=utf-8"
}

start_date = datetime(2023, 11, 24)  # 开始日期
end_date = datetime(2023, 12, 7)    # 结束日期

current_date = start_date
while current_date <= end_date:
    print(current_date.strftime("%Y-%m-%d"))  # 打印当前日期
    startTime = current_date.strftime("%Y-%m-%d")
    if startTime != '2023-12-23':
        current_date += timedelta(days=2)
    print(current_date.strftime("%Y-%m-%d"))  # 打印当前日期
    endTime = current_date.strftime("%Y-%m-%d")
    current_date += timedelta(days=1)  # 增加一天

    response = requests.get(url+'&StartTime='+startTime+'&EndTime='+endTime, headers)
    if response.status_code == 200:
        response.encoding = 'utf-8'
        # 将响应内容转换为 JSON 格式
        json_data = response.json()

        # 循环遍历 JSON 中的数据
        for json_item in json_data['Data']:
            moduleType = json_item['ModuleType']
            if moduleType == 'onepagenews' or moduleType == 'exactnews' or moduleType == 'eventmarketing' or moduleType == 'financeNews' or moduleType == 'notenews':
                ws.append([json_item['UpdateTime'], moduleType, "、".join(
                    classifyItem['ClassifyName'] for classifyItem in json_item['Content']['ClassifyList']),
                           json_item['Content']['Title']])
            elif moduleType == 'shortvideo':
                ws.append([json_item['UpdateTime'], moduleType, "、".join(
                    classifyItem['ClassifyName'] for classifyItem in json_item['Video']['ClassifyList']),
                           json_item['Video']['Title']])
            elif moduleType == 'shorttext':
                ws.append([json_item['UpdateTime'], moduleType, "、".join(
                    classifyItem['ClassifyName'] for classifyItem in json_item['ShortText']['ClassifyList']),
                           json_item['ShortText']['Title']])
            elif moduleType == 'poster':
                ws.append([json_item['UpdateTime'], moduleType, "、".join(
                    classifyItem['ClassifyName'] for classifyItem in json_item['Poster']['ClassifyList']),
                           json_item['Poster']['Name']])

    else:
        print('请求失败，状态码:', response.status_code)

# 保存 Excel 文件
wb.save('example.xlsx')

print('导出成功!!')